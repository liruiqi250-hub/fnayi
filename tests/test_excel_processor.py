from pathlib import Path

from openpyxl import Workbook, load_workbook

from file_toolbox.processors.excel import translate_xlsx
from conftest import FakeTranslator


def test_translate_xlsx_translates_selected_column(tmp_path: Path):
    source = tmp_path / "products.xlsx"
    wb = Workbook()
    ws = wb.active
    ws.append(["ID", "Name"])
    ws.append(["A1", "Product Alpha"])
    wb.save(source)
    wb.close()

    output = translate_xlsx(source, tmp_path / "translated", FakeTranslator(), columns=[2])

    wb_out = load_workbook(output)
    ws_out = wb_out.active
    try:
        assert ws_out["A2"].value == "A1"
        assert ws_out["B2"].value == "[EN] Product Alpha"
    finally:
        wb_out.close()


def test_translate_xlsx_preserves_xlsm_extension(tmp_path: Path):
    source = tmp_path / "products.xlsm"
    wb = Workbook()
    ws = wb.active
    ws.append(["Name"])
    ws.append(["Product Alpha"])
    wb.save(source)
    wb.close()

    output = translate_xlsx(source, tmp_path / "translated", FakeTranslator())

    assert output.suffix == ".xlsm"
