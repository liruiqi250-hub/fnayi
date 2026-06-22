from pathlib import Path
from typing import Callable

from openpyxl import load_workbook

from file_toolbox.translator import Translator


def translate_xlsx(
    source: Path,
    output_dir: Path,
    translator: Translator,
    columns: list[int] | None = None,
    progress_callback: Callable[[int, int], None] | None = None,
) -> Path:
    return translate_xlsx_to_language(
        source, output_dir, translator,
        columns=columns, progress_callback=progress_callback,
    )


def _count_translatable_cells(workbook, selected: set[int]) -> int:
    total = 0
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if selected and cell.column not in selected:
                    continue
                if isinstance(cell.value, str) and cell.value.strip():
                    total += 1
    return total


def translate_xlsx_to_language(
    source: Path,
    output_dir: Path,
    translator: Translator,
    columns: list[int] | None = None,
    target_language: str = "English",
    source_language: str = "Auto-detect",
    target_suffix: str = "EN",
    progress_callback: Callable[[int, int], None] | None = None,
) -> Path:
    output_dir.mkdir(parents=True, exist_ok=True)
    suffix = source.suffix.lower()

    # .xls 格式需要使用 COM 自动化处理
    if suffix == ".csv":
        return translate_csv_to_language(
            source, output_dir, translator,
            columns=columns,
            target_language=target_language,
            source_language=source_language,
            target_suffix=target_suffix,
            progress_callback=progress_callback,
        )
    if suffix == ".xls":
        return _translate_xls_via_com(
            source, output_dir, translator,
            columns=columns,
            target_language=target_language,
            source_language=source_language,
            target_suffix=target_suffix,
            progress_callback=progress_callback,
        )

    workbook = load_workbook(source, keep_vba=True)
    output = output_dir / f"{source.stem}_{target_suffix}{suffix}"

    selected = set(columns or [])
    total = _count_translatable_cells(workbook, selected) if progress_callback else 0

    # Translation cache: avoid re-translating the same cell text
    translation_cache: dict[str, str] = {}

    done = 0
    for sheet in workbook.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                if selected and cell.column not in selected:
                    continue
                if isinstance(cell.value, str) and cell.value.strip():
                    original = cell.value.strip()
                    if original not in translation_cache:
                        translation_cache[original] = translator.translate(
                            original, target_language, source_language
                        )
                    cell.value = translation_cache[original]
                    done += 1
                    if progress_callback:
                        progress_callback(done, total)

    try:
        workbook.save(output)
    finally:
        workbook.close()
    return output


def translate_csv_to_language(
    source: Path,
    output_dir: Path,
    translator: Translator,
    columns: list[int] | None = None,
    target_language: str = "English",
    source_language: str = "Auto-detect",
    target_suffix: str = "EN",
    progress_callback: Callable[[int, int], None] | None = None,
) -> Path:
    """\u7ffb\u8bd1 CSV \u6587\u4ef6\uff0c\u4fdd\u7559 CSV \u683c\u5f0f\u3002"""
    import csv, io
    output_dir.mkdir(parents=True, exist_ok=True)
    output = output_dir / f"{source.stem}_{target_suffix}.csv"

    with open(source, "r", encoding="utf-8-sig", errors="replace") as f:
        reader = csv.reader(f)
        rows = list(reader)

    total = len(rows)
    translation_cache: dict[str, str] = {}
    done = 0

    for row in rows:
        for i, cell in enumerate(row):
            cell = cell.strip()
            if cell:
                if cell not in translation_cache:
                    translation_cache[cell] = translator.translate(
                        cell, target_language, source_language
                    )
                row[i] = translation_cache[cell]
        done += 1
        if progress_callback:
            progress_callback(done, total)

    with open(output, "w", encoding="utf-8-sig", newline="") as f:
        writer = csv.writer(f)
        writer.writerows(rows)
    return output


def _translate_xls_via_com(
    source: Path,
    output_dir: Path,
    translator: Translator,
    columns: list[int] | None = None,
    target_language: str = "English",
    source_language: str = "Auto-detect",
    target_suffix: str = "EN",
    progress_callback: Callable[[int, int], None] | None = None,
) -> Path:
    """Handle legacy .xls via Excel COM automation."""
    import pythoncom
    pythoncom.CoInitialize()
    excel = None
    try:
        from win32com.client import Dispatch
        excel = Dispatch("Excel.Application")
        excel.Visible = False
        excel.DisplayAlerts = False

        wb = excel.Workbooks.Open(str(source.resolve()))
        selected = set(columns or [])

        translation_cache: dict[str, str] = {}
        for sheet in wb.Sheets:
            for row in range(1, sheet.UsedRange.Rows.Count + 1):
                for col in range(1, sheet.UsedRange.Columns.Count + 1):
                    if selected and col not in selected:
                        continue
                    cell = sheet.Cells(row, col)
                    val = cell.Text
                    if val and val.strip():
                        original = val.strip()
                        if original not in translation_cache:
                            translation_cache[original] = translator.translate(
                                original, target_language, source_language
                            )
                        cell.Value = translation_cache[original]

        output = output_dir / f"{source.stem}_{target_suffix}.xlsx"
        wb.SaveAs(str(output.resolve()), FileFormat=51)
        wb.Close(False)
        return output
    except Exception as exc:
        raise RuntimeError(f".xls 处理失败：{exc}")
    finally:
        if excel:
            try: excel.Quit()
            except: pass
        pythoncom.CoUninitialize()


